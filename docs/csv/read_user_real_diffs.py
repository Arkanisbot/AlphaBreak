"""Find REAL differences (not alpha channel noise) between original and user input."""
import openpyxl

user_wb = openpyxl.load_workbook(
    r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-Arkanisbot\AlphaBreak\docs\csv\competitive_feature_matrix_user_input.xlsx",
    data_only=True
)
orig_wb = openpyxl.load_workbook(
    r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-Arkanisbot\AlphaBreak\docs\csv\competitive_feature_matrix.xlsx",
    data_only=True
)

def get_color_stripped(cell):
    """Get color without alpha channel prefix."""
    fill = cell.fill
    for attr in (fill.fgColor, fill.start_color):
        if attr and attr.rgb:
            rgb = str(attr.rgb)
            if rgb not in ("00000000", "0"):
                # Strip leading alpha: FF2F5496 -> 2F5496, 002F5496 -> 2F5496
                return rgb[-6:]
    return "none"

total_diffs = 0
for sheet_name in user_wb.sheetnames:
    if sheet_name == "Legend" or sheet_name not in orig_wb.sheetnames:
        continue
    u_ws = user_wb[sheet_name]
    o_ws = orig_wb[sheet_name]

    for row in range(1, u_ws.max_row + 1):
        for col in range(1, u_ws.max_column + 1):
            u_cell = u_ws.cell(row=row, column=col)
            o_cell = o_ws.cell(row=row, column=col)

            u_val = str(u_cell.value).strip() if u_cell.value is not None else ""
            o_val = str(o_cell.value).strip() if o_cell.value is not None else ""
            u_color = get_color_stripped(u_cell)
            o_color = get_color_stripped(o_cell)

            val_changed = u_val != o_val
            color_changed = u_color != o_color

            if val_changed or color_changed:
                feature = str(u_ws.cell(row=row, column=1).value or "").strip()[:40]
                header = str(u_ws.cell(row=2, column=col).value or "").strip()[:20]
                changes = []
                if val_changed:
                    changes.append(f"Value: '{o_val}' -> '{u_val}'")
                if color_changed:
                    changes.append(f"Color: {o_color} -> {u_color}")
                print(f"  [{sheet_name}] Row {row} Col {col} [{header}] {feature} | {' | '.join(changes)}")
                total_diffs += 1

if total_diffs == 0:
    print("NO REAL DIFFERENCES FOUND (all diffs were alpha channel noise)")
else:
    print(f"\nTotal real differences: {total_diffs}")
