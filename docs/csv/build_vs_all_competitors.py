"""Build head-to-head comparison tabs: each competitor vs AlphaBreak."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
GREEN_FREE = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
GREEN_CHEAP = PatternFill(start_color="60C060", end_color="60C060", fill_type="solid")
GREEN_MID = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
GREEN_PRICEY = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
YELLOW = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
BLUE = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
AB_WIN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
COMP_WIN_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
BOLD_14 = Font(name="Calibri", bold=True, size=14, color="2F5496")
BOLD_12 = Font(name="Calibri", bold=True, size=12)
NORMAL = Font(name="Calibri", size=10)
ITALIC = Font(name="Calibri", size=10, italic=True, color="666666")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"),
)

# AlphaBreak tier → green shade
AB_TIER = {"Free": GREEN_FREE, "Pro": GREEN_MID, "Elite": GREEN_PRICEY, "Planned": BLUE}

# Competitor colors (brand-ish)
COMP_COLORS = {
    "Bloomberg":     PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
    "TradingView":   PatternFill(start_color="2962FF", end_color="2962FF", fill_type="solid"),
    "thinkorswim":   PatternFill(start_color="00897B", end_color="00897B", fill_type="solid"),
    "Seeking Alpha": PatternFill(start_color="FF6D00", end_color="FF6D00", fill_type="solid"),
    "Finviz":        PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
    "Koyfin":        PatternFill(start_color="5C6BC0", end_color="5C6BC0", fill_type="solid"),
    "TrendSpider":   PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid"),
    "StockAnalysis": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "Simply Wall St":PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid"),
    "Barchart":      PatternFill(start_color="455A64", end_color="455A64", fill_type="solid"),
}
COMP_GREENS = {
    "Bloomberg": GREEN_PRICEY, "TradingView": GREEN_CHEAP, "thinkorswim": GREEN_FREE,
    "Seeking Alpha": GREEN_CHEAP, "Finviz": GREEN_CHEAP, "Koyfin": GREEN_CHEAP,
    "TrendSpider": GREEN_MID, "StockAnalysis": GREEN_FREE, "Simply Wall St": GREEN_CHEAP,
    "Barchart": GREEN_MID,
}
AB_GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

# ── Competitor data ─────────────────────────────────────────────────────────
COMPETITORS = {
    "Bloomberg": {
        "price": "$24,000/yr (~$2,000/mo)",
        "type": "Institutional terminal",
        "strengths": "Unmatched data depth, supply chain, real-time, BQL scripting",
    },
    "TradingView": {
        "price": "Free–$60/mo",
        "type": "Charting / TA platform",
        "strengths": "Best-in-class charting UX, Pine Script, huge community",
    },
    "thinkorswim": {
        "price": "Free (with Schwab account)",
        "type": "Brokerage / options platform",
        "strengths": "Best free options tools, integrated trading, ThinkScript",
    },
    "Seeking Alpha": {
        "price": "Free–$42/mo",
        "type": "Fundamental / quant ratings",
        "strengths": "Best quant scoring (A-F grades), dividend analysis, crowd-sourced articles",
    },
    "Finviz": {
        "price": "Free–$40/mo",
        "type": "Screener / snapshot tool",
        "strengths": "Best screener/heatmap, 70-metric snapshot, fast, pattern recognition",
    },
    "Koyfin": {
        "price": "Free–$45/mo",
        "type": "Financial data terminal",
        "strengths": "Bloomberg-lite fundamentals, historical valuation charts, segment data",
    },
    "TrendSpider": {
        "price": "$32–$87/mo (no free tier)",
        "type": "Automated technical analysis",
        "strengths": "Auto-trendlines, auto-Fib, Raindrop charts, multi-TF, smart checklists",
    },
    "StockAnalysis": {
        "price": "Free–$10/mo",
        "type": "Free fundamental data",
        "strengths": "Best free fundamentals, 10yr+ financials, segment data, clean UI",
    },
    "Simply Wall St": {
        "price": "Free–$40/mo",
        "type": "Visual fundamental analysis",
        "strengths": "Snowflake ratings, DCF model, visual reports, beginner-friendly",
    },
    "Barchart": {
        "price": "Free–$100/mo",
        "type": "Options + technical data",
        "strengths": "Good options data, unusual activity screener, commodities/futures depth",
    },
}

# ── Feature comparison data per competitor ──────────────────────────────────
# Each entry: (feature, comp_status, ab_status, verdict)
# Verdict keywords: "ADVANTAGE: AlphaBreak", "ADVANTAGE: {comp}", "Parity", or descriptive

def make_features(comp):
    """Return list of (section, [(feature, comp_status, ab_status, verdict)])."""
    C = comp
    data = {
        "Bloomberg": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements (10yr+)",    "10yr+", "Free",    "AlphaBreak needs deeper history"),
                ("Revenue by segment / geography",  "Yes",   "Elite",   "AlphaBreak at $299 vs $2K/mo"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "Yes",   "Free",    "Parity"),
                ("Earnings call transcripts",       "Yes",   "Planned", "Bloomberg + Seeking Alpha only"),
                ("Peer comparison table",           "Best",  "Pro",     "Bloomberg COMP is gold standard"),
                ("Dividend analysis",               "Yes",   "Pro",     "Parity at Pro tier"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity at Pro tier"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "AlphaBreak: 20 funds, 8.4M rows"),
                ("Supply chain mapping",            "SPLC",  "Elite",   "Bloomberg-exclusive, AB plans it"),
                ("3-statement financials w/ trends","Yes",   "Elite",   "AlphaBreak at 1/80th the price"),
                ("DuPont decomposition",            "Yes",   "Elite",   "AlphaBreak at 1/80th the price"),
                ("DCF / intrinsic value model",     "Yes",   "Planned", "Bloomberg has full modeling suite"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Yes",   "Free",    "Parity"),
                ("Technical indicators",            "100+",  "Free",    "Bloomberg more; AB covers essentials"),
                ("Drawing tools",                   "Yes",   "Pro", "Bloomberg has full suite"),
                ("Multi-timeframe analysis",        "Yes",   "Pro",     "Parity at Pro tier"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Auto Fibonacci levels",           "No",    "No",   "Neither offers"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Custom scripting (BQL)",          "BQL",   "No",      "ADVANTAGE: Bloomberg"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "OMON",  "Free",    "Bloomberg more depth; AB free"),
                ("Fair value (Black-Scholes)",      "OVME",  "Pro",    "Both offer; AB free"),
                ("IV history / vol surface",        "SKEW",  "Elite",   "Bloomberg best-in-class"),
                ("Unusual options activity",        "Yes",   "Pro",     "AB at 1/240th the price"),
                ("Probability of profit",           "Yes",   "Pro",   "Parity at Elite tier"),
                ("Strategy builder + P&L diagram",  "Yes",   "Planned", "Bloomberg has full suite"),
                ("Market Maker Move",               "Yes",   "Pro",   "Parity at Elite tier"),
                ("IV crush modeling",               "Yes",   "Elite",   "Parity at Elite tier"),
                ("ORATS historical data",           "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "Yes",   "Elite",   "AB at 1/80th the price"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("News NLP sentiment scoring",      "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Earnings surprise prediction",    "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Short interest / squeeze score",  "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "PORT",  "Free",    "Bloomberg live trading; AB paper"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade thesis builder",            "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Pre/post-trade reviews",          "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Community / shared journal",      "IB Chat","Free",   "Different approach — both social"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "Yes",   "Elite",   "Bloomberg more sophisticated"),
                ("Brokerage integration",           "Yes",   "Planned", "Bloomberg is the brokerage"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "Yes",   "Pro",   "Bloomberg included; AB $299/mo"),
                ("API access",                      "BQL",   "API tier","Bloomberg BQL more powerful"),
                ("WebSocket streaming",             "Yes",   "Pro",   "Parity at Elite tier"),
                ("Mobile app",                      "Yes",   "Planned", "Bloomberg has mature mobile"),
                ("Dark pool data",                  "Yes",   "Pro",    "Both offer; AB free (621K rows)"),
                ("Forex correlations",              "Yes",   "Free",    "Both offer; AB 21 pairs, 54yr"),
            ]),
        ],
        "TradingView": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements",            "Basic", "Free",    "ADVANTAGE: AlphaBreak — deeper data"),
                ("Revenue by segment / geography",  "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Earnings call transcripts",       "No",    "Planned", "Neither offers yet"),
                ("Peer comparison table",           "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Dividend analysis",               "Basic", "Pro",     "ADVANTAGE: AlphaBreak — deeper"),
                ("Insider trading data",            "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Institutional ownership / 13F",   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Supply chain mapping",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("3-statement financials w/ trends","No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DCF / intrinsic value model",     "No",    "Planned", "Neither offers yet"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Best",  "Free",    "ADVANTAGE: TradingView — best UX"),
                ("Technical indicators",            "100+",  "Free",    "ADVANTAGE: TradingView — more depth"),
                ("Drawing tools",                   "Best",  "Pro", "ADVANTAGE: TradingView"),
                ("Multi-timeframe analysis",        "Yes",   "Pro",     "Parity"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Auto Fibonacci levels",           "No",    "No",   "Neither offers"),
                ("Seasonality heatmap",             "Community","Pro",  "ADVANTAGE: AlphaBreak — native"),
                ("Custom scripting",                "Pine Script","No", "ADVANTAGE: TradingView"),
                ("Technical opinion signal",        "Widget","Free",    "Parity"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Unusual options activity",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Probability of profit",           "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Strategy builder + P&L diagram",  "No",    "Planned", "Neither yet; AB plans it"),
                ("Market Maker Move",               "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("IV crush modeling",               "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("News NLP sentiment scoring",      "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Earnings surprise prediction",    "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Short interest / squeeze score",  "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade thesis builder",            "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Pre/post-trade reviews",          "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Community / shared journal",      "Ideas", "Free",    "TV Ideas is strong; AB has journal"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "Yes",   "Free",    "Parity — TV alerts are excellent"),
                ("Backtesting",                     "Pine",  "Elite",   "TradingView Pine backtests are strong"),
                ("Brokerage integration",           "Multi-broker","Planned","ADVANTAGE: TradingView"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "Paid",  "Pro",   "Both paid tiers"),
                ("API access",                      "No",    "API tier","ADVANTAGE: AlphaBreak"),
                ("WebSocket streaming",             "Yes",   "Pro",   "Parity"),
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: TradingView"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "Yes",   "Free",    "Parity"),
            ]),
        ],
        "thinkorswim": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements",            "5yr",   "Free",    "ADVANTAGE: AlphaBreak — needs 10yr+"),
                ("Revenue by segment / geography",  "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "Yes",   "Free",    "Parity"),
                ("Earnings call transcripts",       "No",    "Planned", "Neither offers yet"),
                ("Peer comparison table",           "Basic", "Pro",     "ADVANTAGE: AlphaBreak — deeper"),
                ("Dividend analysis",               "Basic", "Pro",     "ADVANTAGE: AlphaBreak — deeper"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity at Pro tier"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity; AB has deeper archive"),
                ("Supply chain mapping",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("3-statement financials w/ trends","No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DCF / intrinsic value model",     "No",    "Planned", "Neither offers yet"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Yes",   "Free",    "Parity — ToS has depth"),
                ("Technical indicators",            "400+",  "Free",    "ADVANTAGE: thinkorswim — 400+ studies"),
                ("Drawing tools",                   "Yes",   "Pro", "ADVANTAGE: thinkorswim"),
                ("Multi-timeframe analysis",        "Yes",   "Pro",     "Parity"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Auto Fibonacci levels",           "No",    "No",   "Neither offers"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Custom scripting",                "ThinkScript","No", "ADVANTAGE: thinkorswim"),
                ("Technical opinion signal",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "Best",  "Free",    "ADVANTAGE: thinkorswim — industry best"),
                ("Fair value (Black-Scholes)",      "Yes",   "Pro",    "Parity — both offer"),
                ("IV history / vol surface",        "Yes",   "Elite",   "ADVANTAGE: thinkorswim — free"),
                ("Unusual options activity",        "Sizzle","Pro",     "ToS Sizzle is well-known; AB Pro"),
                ("Probability of profit",           "Yes",   "Pro",   "ADVANTAGE: thinkorswim — free"),
                ("Strategy builder + P&L diagram",  "Best",  "Planned", "ADVANTAGE: thinkorswim — best in class"),
                ("Market Maker Move",               "Yes",   "Pro",   "ToS free; AB Elite"),
                ("IV crush modeling",               "Partial","Elite",  "Parity — both have it"),
                ("ORATS historical data",           "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("News NLP sentiment scoring",      "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Earnings surprise prediction",    "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Short interest / squeeze score",  "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "Live",  "Free",    "ADVANTAGE: thinkorswim — live trading"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade thesis builder",            "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Pre/post-trade reviews",          "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Community / shared journal",      "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "Yes",   "Elite",   "ToS free backtesting; AB at Elite"),
                ("Brokerage integration",           "Schwab","Planned", "ADVANTAGE: thinkorswim — IS the broker"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "Yes",   "Pro",   "ADVANTAGE: thinkorswim — free"),
                ("API access",                      "No",    "API tier","ADVANTAGE: AlphaBreak"),
                ("WebSocket streaming",             "Yes",   "Pro",   "ToS free; AB Elite"),
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: thinkorswim"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "Yes",   "Free",    "Parity"),
            ]),
        ],
        "Seeking Alpha": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements",            "Yes",   "Free",    "Parity"),
                ("Revenue by segment / geography",  "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "Best",  "Free",    "ADVANTAGE: Seeking Alpha — best revisions"),
                ("Earnings call transcripts",       "Yes",   "Planned", "ADVANTAGE: Seeking Alpha"),
                ("Peer comparison table",           "Yes",   "Pro",     "Parity at Pro tier"),
                ("Dividend analysis",               "Best",  "Pro",     "ADVANTAGE: Seeking Alpha — best dividends"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity at Pro tier"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity; AB has deeper archive"),
                ("Supply chain mapping",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("3-statement financials w/ trends","No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DCF / intrinsic value model",     "No",    "Planned", "Neither offers yet"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Technical indicators",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Drawing tools",                   "No",    "Pro", "ADVANTAGE: AlphaBreak (planned)"),
                ("Multi-timeframe analysis",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Auto Fibonacci levels",           "No",    "No",   "Neither offers"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Custom scripting",                "No",    "No",      "Neither offers"),
                ("Technical opinion signal",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Unusual options activity",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Probability of profit",           "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Strategy builder + P&L diagram",  "No",    "Planned", "Neither yet"),
                ("Market Maker Move",               "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("IV crush modeling",               "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "Yes",   "Pro",     "SA pioneered this; AB matches at Pro"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("News NLP sentiment scoring",      "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Earnings surprise prediction",    "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "Partial","Free",   "ADVANTAGE: AlphaBreak — more comprehensive"),
                ("Short interest / squeeze score",  "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "Basic", "Free",    "ADVANTAGE: AlphaBreak — more features"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade thesis builder",            "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Pre/post-trade reviews",          "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Community / shared journal",      "Articles","Free",  "SA has articles; AB has journal"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Brokerage integration",           "No",    "Planned", "Neither yet"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("API access",                      "No",    "API tier","ADVANTAGE: AlphaBreak"),
                ("WebSocket streaming",             "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: Seeking Alpha"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
        ],
        "Finviz": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "70 metrics","Free","ADVANTAGE: Finviz — best snapshot density"),
                ("Financial statements",            "Snapshot","Free",  "ADVANTAGE: AlphaBreak — full history"),
                ("Revenue by segment / geography",  "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Earnings call transcripts",       "No",    "Planned", "Neither offers"),
                ("Peer comparison table",           "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Dividend analysis",               "Basic", "Pro",     "ADVANTAGE: AlphaBreak — deeper"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity at Pro tier"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity; AB has deeper archive"),
                ("Supply chain mapping",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("3-statement financials w/ trends","No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DCF / intrinsic value model",     "No",    "Planned", "Neither offers"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Basic", "Free",    "ADVANTAGE: AlphaBreak"),
                ("Technical indicators",            "Basic", "Free",    "ADVANTAGE: AlphaBreak"),
                ("Drawing tools",                   "No",    "Pro", "ADVANTAGE: AlphaBreak (planned)"),
                ("Multi-timeframe analysis",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Auto Fibonacci levels",           "No",    "No",   "Neither offers"),
                ("Candlestick pattern recognition", "Yes",   "Planned", "ADVANTAGE: Finviz — has it now"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Technical opinion signal",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Unusual options activity",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Probability of profit",           "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Market Maker Move",               "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Short interest / squeeze score",  "Yes",   "Pro",     "Finviz free; AB at Pro tier"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Backtesting",                     "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "Paid",  "Pro",   "Both paid tiers"),
                ("Screener / heatmap",              "Best",  "No",      "ADVANTAGE: Finviz — iconic screener"),
                ("Mobile app",                      "No",    "Planned", "Neither yet"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
        ],
        "Koyfin": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements (10yr+)",    "10yr+", "Free",    "Parity — both deep"),
                ("Revenue by segment / geography",  "Yes",   "Elite",   "Koyfin at $45; AB at $299"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "Yes",   "Free",    "Parity"),
                ("Earnings call transcripts",       "No",    "Planned", "Neither offers"),
                ("Peer comparison table",           "Yes",   "Pro",     "Parity — both strong"),
                ("Dividend analysis",               "Yes",   "Pro",     "Parity"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity; AB has deeper archive"),
                ("Supply chain mapping",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Historical valuation time series","Best",  "No",      "ADVANTAGE: Koyfin — unique feature"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DCF / intrinsic value model",     "No",    "Planned", "Neither offers"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Yes",   "Free",    "Parity"),
                ("Technical indicators",            "Yes",   "Free",    "Parity"),
                ("Drawing tools",                   "Basic", "Pro", "Parity — both limited"),
                ("Multi-timeframe analysis",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Auto Fibonacci levels",           "No",    "No",   "Neither offers"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Technical opinion signal",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Unusual options activity",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Probability of profit",           "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Market Maker Move",               "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Short interest / squeeze score",  "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "Yes",   "Free",    "Parity"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Backtesting",                     "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("Data & Infrastructure", [
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: Koyfin"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
        ],
        "TrendSpider": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Financial statements",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Earnings estimates & revisions",  "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Peer comparison table",           "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Dividend analysis",               "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Insider trading data",            "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Institutional ownership / 13F",   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Yes",   "Free",    "Parity"),
                ("Technical indicators",            "Yes",   "Free",    "Parity"),
                ("Drawing tools",                   "Yes",   "Pro", "ADVANTAGE: TrendSpider"),
                ("Multi-timeframe analysis",        "Best",  "Pro",     "ADVANTAGE: TrendSpider — best MTF"),
                ("Auto-detected trendlines",        "Yes",   "Pro",   "TrendSpider pioneered this; AB plans it"),
                ("Auto Fibonacci levels",           "Yes",   "No",   "TrendSpider pioneered this; AB plans it"),
                ("Candlestick pattern recognition", "Best",  "Planned", "ADVANTAGE: TrendSpider"),
                ("Seasonality heatmap",             "Yes",   "Pro",     "Parity"),
                ("Raindrop / volume profile",       "Yes",   "Planned", "ADVANTAGE: TrendSpider — proprietary"),
                ("Smart checklists",                "Yes",   "Pro",     "Parity"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "Partial","Elite",  "ADVANTAGE: AlphaBreak — deeper"),
                ("Unusual options activity",        "Partial","Pro",    "ADVANTAGE: AlphaBreak — deeper"),
                ("Options flow dashboard",          "Partial","Elite",  "ADVANTAGE: AlphaBreak — deeper"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("News NLP sentiment scoring",      "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "Yes",   "Elite",   "Both offer; TS code-free builder"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "Paid",  "Pro",   "Both paid tiers"),
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: TrendSpider"),
                ("Dark pool data",                  "Partial","Pro",   "ADVANTAGE: AlphaBreak — 621K rows free"),
            ]),
        ],
        "StockAnalysis": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements (10yr+)",    "10yr+", "Free",    "Parity — both excellent free"),
                ("Revenue by segment / geography",  "Yes",   "Elite",   "SA free; AB at Elite — ADVANTAGE: StockAnalysis"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "Yes",   "Free",    "Parity"),
                ("Peer comparison table",           "Limited","Pro",    "ADVANTAGE: AlphaBreak — deeper"),
                ("Dividend analysis",               "Yes",   "Pro",     "Parity"),
                ("Insider trading data",            "Yes",   "Pro",     "SA free; AB at Pro — ADVANTAGE: StockAnalysis"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity; AB has deeper archive"),
                ("3-statement financials w/ trends","Yes",   "Elite",   "SA free; AB at Elite — ADVANTAGE: StockAnalysis"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DCF / intrinsic value model",     "No",    "Planned", "Neither offers"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Technical indicators",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Multi-timeframe analysis",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Technical opinion signal",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Options Analytics", [
                ("Options chain",                   "Basic", "Free",    "ADVANTAGE: AlphaBreak — Greeks, fair value"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Unusual options activity",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
        ],
        "Simply Wall St": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements",            "Yes",   "Free",    "Parity"),
                ("Revenue by segment / geography",  "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Peer comparison table",           "Yes",   "Pro",     "Parity"),
                ("Dividend analysis",               "Yes",   "Pro",     "Parity"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity"),
                ("DCF / intrinsic value model",     "Yes",   "Planned", "ADVANTAGE: Simply Wall St — built in"),
                ("Snowflake 5-factor visual",       "Yes",   "No",      "ADVANTAGE: Simply Wall St — unique"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Technical indicators",            "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Multi-timeframe analysis",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Technical opinion signal",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Unusual options activity",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "Snowflake","Pro",  "Both have scoring; different approach"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "Yes",   "Free",    "Parity"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("Data & Infrastructure", [
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: Simply Wall St"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
        ],
        "Barchart": [
            ("Overview & Fundamentals", [
                ("Company overview / key stats",    "Yes",   "Free",    "Parity"),
                ("Financial statements",            "Yes",   "Free",    "Parity"),
                ("Revenue by segment / geography",  "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("Analyst consensus & targets",     "Yes",   "Free",    "Parity"),
                ("Earnings estimates & revisions",  "Yes",   "Free",    "Parity"),
                ("Peer comparison table",           "Basic", "Pro",     "ADVANTAGE: AlphaBreak — deeper"),
                ("Dividend analysis",               "Basic", "Pro",     "ADVANTAGE: AlphaBreak — deeper"),
                ("Insider trading data",            "Yes",   "Pro",     "Parity at Pro tier"),
                ("Institutional ownership / 13F",   "Yes",   "Free",    "Parity"),
                ("Supply chain mapping",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
                ("DuPont decomposition",            "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("Charting & Technical Analysis", [
                ("Interactive charting",            "Yes",   "Free",    "Parity"),
                ("Technical indicators",            "Yes",   "Free",    "Parity"),
                ("Drawing tools",                   "Yes",   "Pro", "ADVANTAGE: Barchart"),
                ("Multi-timeframe analysis",        "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Auto-detected trendlines",        "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Seasonality heatmap",             "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Technical opinion signal",        "Yes",   "Free",    "Parity"),
            ]),
            ("Options Analytics", [
                ("Options chain + Greeks",          "Yes",   "Free",    "Parity"),
                ("Fair value (Black-Scholes)",      "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("IV history / vol surface",        "Partial","Elite",  "ADVANTAGE: AlphaBreak — deeper"),
                ("Unusual options activity",        "Yes",   "Pro",     "Barchart free; AB at Pro"),
                ("Probability of profit",           "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Market Maker Move",               "No",    "Pro",   "ADVANTAGE: AlphaBreak"),
                ("Options flow dashboard",          "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("AI, Scoring & Quantitative", [
                ("AI trade scoring",                "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Composite tech+fundamental score","No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Quant letter grades (A-F)",       "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Smart checklists",                "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
                ("Regime-aware analysis",           "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Signal accuracy tracking",        "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Short interest / squeeze score",  "No",    "Pro",     "ADVANTAGE: AlphaBreak"),
            ]),
            ("Portfolio, Journal & Workflow", [
                ("Portfolio tracker",               "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Trade journal",                   "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("AI-scored journal entries",       "No",    "Free",    "ADVANTAGE: AlphaBreak"),
                ("Watchlist",                       "Yes",   "Free",    "Parity"),
                ("Email / push notifications",      "Yes",   "Free",    "Parity"),
                ("Backtesting",                     "No",    "Elite",   "ADVANTAGE: AlphaBreak"),
            ]),
            ("Data & Infrastructure", [
                ("Real-time data",                  "Paid",  "Pro",   "Both paid tiers"),
                ("Mobile app",                      "Yes",   "Planned", "ADVANTAGE: Barchart"),
                ("Dark pool data",                  "No",    "Pro",    "ADVANTAGE: AlphaBreak"),
                ("Forex correlations",              "No",    "Free",    "ADVANTAGE: AlphaBreak"),
            ]),
        ],
    }
    return data[comp]


def status_fill(status, comp_name):
    """Return (fill, text, white_font)."""
    s = status.strip()
    if s in ("No",):
        return RED, "No", True
    if s == "Planned":
        return BLUE, "Planned", False
    if s in ("Partial", "Basic", "Snapshot", "Limited", "Community", "Sizzle",
             "Custom scan", "Snowflake", "Widget"):
        return YELLOW, s, False
    # Some form of "yes"
    if comp_name == "AlphaBreak":
        for tier, fill in AB_TIER.items():
            if s == tier:
                return fill, s, False
        return GREEN_FREE, s, False
    green = COMP_GREENS.get(comp_name, GREEN_MID)
    return green, s, False


def style_cell(cell, fill, text, white_font=False):
    cell.value = text
    cell.fill = fill
    cell.font = Font(name="Calibri", size=10, color="FFFFFF" if white_font else "000000")
    cell.alignment = CENTER
    cell.border = THIN


def build_vs_tab(ws, comp_name, comp_info, features_data):
    """Build a competitor vs AlphaBreak tab."""
    comp_fill = COMP_COLORS.get(comp_name, HEADER_FILL)
    price = comp_info["price"]
    ctype = comp_info["type"]
    strengths = comp_info["strengths"]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1,
        value=f"{comp_name} ({price}) vs AlphaBreak (Free–$299/mo)").font = BOLD_14
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(row=2, column=1,
        value=f"{comp_name}: {ctype} | Strengths: {strengths}").font = ITALIC
    ws.row_dimensions[2].height = 24

    # Headers
    row = 3
    for col, (hdr, fill) in enumerate([
        ("Feature", HEADER_FILL), (comp_name, comp_fill),
        ("AlphaBreak", AB_GREEN), ("Verdict", HEADER_FILL)
    ], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = fill
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN
    ws.row_dimensions[3].height = 28

    # Counters
    ab_wins = 0
    comp_wins = 0
    parity_count = 0

    row = 4
    for section_name, features in features_data:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        sc = ws.cell(row=row, column=1, value=section_name)
        sc.fill = SECTION_FILL
        sc.font = BOLD
        sc.alignment = LEFT
        sc.border = THIN
        ws.row_dimensions[row].height = 24
        row += 1

        for feat, comp_status, ab_status, verdict in features:
            fc = ws.cell(row=row, column=1, value=feat)
            fc.font = BOLD
            fc.alignment = LEFT
            fc.border = THIN

            cf, ct, cw = status_fill(comp_status, comp_name)
            style_cell(ws.cell(row=row, column=2), cf, ct, cw)

            af, at_, aw = status_fill(ab_status, "AlphaBreak")
            style_cell(ws.cell(row=row, column=3), af, at_, aw)

            vc = ws.cell(row=row, column=4, value=verdict)
            vc.alignment = LEFT
            vc.border = THIN
            if "ADVANTAGE: AlphaBreak" in verdict:
                vc.fill = AB_WIN_FILL
                vc.font = Font(name="Calibri", size=10, bold=True, color="375623")
                ab_wins += 1
            elif f"ADVANTAGE: {comp_name}" in verdict:
                vc.fill = COMP_WIN_FILL
                vc.font = Font(name="Calibri", size=10, bold=True, color="833C0C")
                comp_wins += 1
            elif "Parity" in verdict or "Both" in verdict or "Neither" in verdict:
                vc.font = NORMAL
                parity_count += 1
            else:
                vc.font = NORMAL

            ws.row_dimensions[row].height = 22
            row += 1
        row += 1

    # Scorecard
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="SCORECARD").font = BOLD_14
    ws.row_dimensions[row].height = 28
    row += 1

    total = ab_wins + comp_wins + parity_count
    for label, value, fill in [
        ("AlphaBreak Advantages", str(ab_wins), AB_WIN_FILL),
        (f"{comp_name} Advantages", str(comp_wins), COMP_WIN_FILL),
        ("Parity / Comparable", str(parity_count), NEUTRAL_FILL),
        ("Total Features Compared", str(total), NEUTRAL_FILL),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=1).border = THIN
        ws.cell(row=row, column=1).fill = fill
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = BOLD_12
        vc.alignment = CENTER
        vc.border = THIN
        vc.fill = fill
        ws.row_dimensions[row].height = 24
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 46

    ws.freeze_panes = "A4"

    return ab_wins, comp_wins, parity_count


# ── Build all tabs ──────────────────────────────────────────────────────────
all_scores = {}
first = True
for comp_name, comp_info in COMPETITORS.items():
    if first:
        ws = wb.active
        ws.title = f"vs {comp_name}"[:31]
        first = False
    else:
        ws = wb.create_sheet(f"vs {comp_name}"[:31])
    features_data = make_features(comp_name)
    ab_w, comp_w, par = build_vs_tab(ws, comp_name, comp_info, features_data)
    all_scores[comp_name] = (ab_w, comp_w, par)


# ── Summary tab ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Summary")
ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
ws_sum.cell(row=1, column=1, value="AlphaBreak vs All Competitors — Scorecard Summary").font = BOLD_14
ws_sum.row_dimensions[1].height = 30

headers = ["Competitor", "Price", "AB Wins", "Comp Wins", "Parity", "AB Win %"]
for col, h in enumerate(headers, 1):
    c = ws_sum.cell(row=2, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = THIN
ws_sum.row_dimensions[2].height = 28

row = 3
for comp_name, (ab_w, comp_w, par) in all_scores.items():
    total = ab_w + comp_w + par
    win_pct = f"{ab_w / total * 100:.0f}%" if total > 0 else "N/A"

    ws_sum.cell(row=row, column=1, value=comp_name).font = BOLD
    ws_sum.cell(row=row, column=1).border = THIN
    ws_sum.cell(row=row, column=2, value=COMPETITORS[comp_name]["price"]).font = NORMAL
    ws_sum.cell(row=row, column=2).alignment = CENTER
    ws_sum.cell(row=row, column=2).border = THIN

    ac = ws_sum.cell(row=row, column=3, value=ab_w)
    ac.fill = AB_WIN_FILL
    ac.font = BOLD_12
    ac.alignment = CENTER
    ac.border = THIN

    cc = ws_sum.cell(row=row, column=4, value=comp_w)
    cc.fill = COMP_WIN_FILL
    cc.font = BOLD_12
    cc.alignment = CENTER
    cc.border = THIN

    pc = ws_sum.cell(row=row, column=5, value=par)
    pc.fill = NEUTRAL_FILL
    pc.font = BOLD_12
    pc.alignment = CENTER
    pc.border = THIN

    wc = ws_sum.cell(row=row, column=6, value=win_pct)
    wc.font = Font(name="Calibri", bold=True, size=12, color="375623")
    wc.fill = AB_WIN_FILL
    wc.alignment = CENTER
    wc.border = THIN

    ws_sum.row_dimensions[row].height = 24
    row += 1

ws_sum.column_dimensions["A"].width = 20
ws_sum.column_dimensions["B"].width = 28
ws_sum.column_dimensions["C"].width = 14
ws_sum.column_dimensions["D"].width = 14
ws_sum.column_dimensions["E"].width = 14
ws_sum.column_dimensions["F"].width = 14
ws_sum.freeze_panes = "A3"


# ── Legend tab ──────────────────────────────────────────────────────────────
ws_leg = wb.create_sheet("Legend")
ws_leg.column_dimensions["A"].width = 22
ws_leg.column_dimensions["B"].width = 55
legend = [
    ("Color Code", "Meaning", None),
    ("Dark Green", "Free / included with free account", GREEN_FREE),
    ("Medium Green", "Available for <$50/mo", GREEN_CHEAP),
    ("Light Green", "Available for $50-150/mo", GREEN_MID),
    ("Pale Green", "Available for $150+/mo or institutional", GREEN_PRICEY),
    ("Red", "Feature NOT offered", RED),
    ("Yellow", "Partial / basic implementation", YELLOW),
    ("Blue", "Planned for AlphaBreak (not yet built)", BLUE),
    ("", "", None),
    ("Verdict Colors", "", None),
    ("Green background", "AlphaBreak advantage", AB_WIN_FILL),
    ("Orange background", "Competitor advantage", COMP_WIN_FILL),
    ("Gray background", "Parity / comparable", NEUTRAL_FILL),
]
for r, (label, desc, fill) in enumerate(legend, 1):
    a = ws_leg.cell(row=r, column=1, value=label)
    b = ws_leg.cell(row=r, column=2, value=desc)
    if r == 1 or (label and label.endswith("Colors")):
        a.font = BOLD
        b.font = BOLD
    elif fill:
        a.fill = fill
        needs_white = fill == RED
        a.font = Font(name="Calibri", size=11, color="FFFFFF" if needs_white else "000000")


output = r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-SophistryDude\AlphaBreak\docs\csv\competitive_vs_alphabreak.xlsx"
wb.save(output)
print(f"Saved to {output}")
print("\nScorecard Summary:")
for comp, (ab_w, comp_w, par) in all_scores.items():
    total = ab_w + comp_w + par
    pct = ab_w / total * 100 if total else 0
    print(f"  vs {comp:20s}  AB: {ab_w:2d}  Comp: {comp_w:2d}  Parity: {par:2d}  AB Win: {pct:.0f}%")
