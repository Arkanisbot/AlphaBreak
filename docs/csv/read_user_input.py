"""Read the user's updated AlphaBreak column from their modified xlsx."""
import openpyxl
import json

wb = openpyxl.load_workbook(
    r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-SophistryDude\AlphaBreak\docs\csv\competitive_feature_matrix_user_input.xlsx",
    data_only=True
)

# For each sheet, read the AlphaBreak column (last data column) and capture
# the cell value + fill color for every feature row
results = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if sheet_name == "Legend":
        continue

    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        # Get feature name (col A) and AlphaBreak cell (last col)
        feature_cell = row[0]
        ab_cell = row[-1]  # Last column = AlphaBreak

        if feature_cell.value and ab_cell.value is not None:
            # Get fill color
            fill = ab_cell.fill
            fg_color = "none"
            if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                fg_color = str(fill.fgColor.rgb)
            elif fill and fill.start_color and fill.start_color.rgb and fill.start_color.rgb != "00000000":
                fg_color = str(fill.start_color.rgb)

            rows_data.append({
                "feature": str(feature_cell.value).strip(),
                "ab_value": str(ab_cell.value).strip() if ab_cell.value else "",
                "ab_fill": fg_color,
                "row_num": feature_cell.row,
            })

    results[sheet_name] = rows_data

# Print everything
for sheet_name, rows in results.items():
    print(f"\n=== {sheet_name} ===")
    for r in rows:
        print(f"  Row {r['row_num']:3d} | {r['feature']:45s} | Value: {r['ab_value']:20s} | Fill: {r['ab_fill']}")
