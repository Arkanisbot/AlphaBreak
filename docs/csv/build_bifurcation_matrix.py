"""Build bifurcation analysis + Bloomberg vs AlphaBreak comparison."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
GREEN_FREE = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
GREEN_CHEAP = PatternFill(start_color="60C060", end_color="60C060", fill_type="solid")
GREEN_MID = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
GREEN_PRICEY = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
YELLOW = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
BLUE = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CHARTING_HEADER = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Purple for TA side
FUNDAMENTAL_HEADER = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")  # Blue for fundamental side
BOTH_HEADER = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")  # Gold for both
NEITHER_HEADER = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray for neither
BB_BLUE = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")  # Bloomberg orange
AB_GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # AlphaBreak green

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
BOLD_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD_14 = Font(name="Calibri", bold=True, size=14, color="2F5496")
BOLD_12 = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
NORMAL = Font(name="Calibri", size=10)
ITALIC = Font(name="Calibri", size=10, italic=True, color="666666")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"),
)

# ── Bifurcation data ───────────────────────────────────────────────────────
# Competitors grouped: Charting/TA-first vs Fundamental-first
# Charting side: TradingView, thinkorswim, TrendSpider, Barchart
# Fundamental side: Seeking Alpha, Koyfin, StockAnalysis, Simply Wall St, Finviz
# Both: Bloomberg (and AlphaBreak aims to be here)

CHARTING_COMPS = [
    ("TradingView", "Free–$60/mo"),
    ("thinkorswim", "Free"),
    ("TrendSpider", "$32–87/mo"),
    ("Barchart", "Free–$100/mo"),
]
FUNDAMENTAL_COMPS = [
    ("Seeking Alpha", "Free–$42/mo"),
    ("Finviz", "Free–$40/mo"),
    ("Koyfin", "Free–$45/mo"),
    ("StockAnalysis", "Free–$10/mo"),
    ("Simply Wall St", "Free–$40/mo"),
]
BRIDGE_COMPS = [
    ("Bloomberg", "$24,000/yr"),
    ("AlphaBreak", "Free / $99 / $299"),
]

# Green shade per competitor
COMP_GREENS = {
    "TradingView": GREEN_CHEAP, "thinkorswim": GREEN_FREE, "TrendSpider": GREEN_MID,
    "Barchart": GREEN_MID, "Seeking Alpha": GREEN_CHEAP, "Finviz": GREEN_CHEAP,
    "Koyfin": GREEN_CHEAP, "StockAnalysis": GREEN_FREE, "Simply Wall St": GREEN_CHEAP,
    "Bloomberg": GREEN_PRICEY, "AlphaBreak": GREEN_FREE,
}

# Features grouped by which side of the bifurcation they fall on
# "charting" = features the charting/TA platforms excel at
# "fundamental" = features the fundamental platforms excel at
# "options" = options analytics (mostly charting-side but sparse)
# "ai_quant" = AI/scoring (almost nobody)
# "journal" = portfolio/journal/workflow (almost nobody)

BIFURCATION = {
    "Charting & Technical Analysis Features": {
        "header_fill": CHARTING_HEADER,
        "note": "These features are dominated by TradingView, thinkorswim, and TrendSpider. Fundamental platforms largely ignore them.",
        "features": [
            # (feature, {competitor: status})
            ("Interactive charting",
             {"TradingView": "Best", "thinkorswim": "Yes", "TrendSpider": "Yes", "Barchart": "Yes",
              "Seeking Alpha": "No", "Finviz": "Basic", "Koyfin": "Yes", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Free"}),
            ("Technical indicators (100+)",
             {"TradingView": "100+", "thinkorswim": "400+", "TrendSpider": "Yes", "Barchart": "Yes",
              "Seeking Alpha": "No", "Finviz": "Basic", "Koyfin": "Yes", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "100+", "AlphaBreak": "Free"}),
            ("Drawing tools",
             {"TradingView": "Best", "thinkorswim": "Yes", "TrendSpider": "Yes", "Barchart": "Yes",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "Basic", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Multi-timeframe analysis",
             {"TradingView": "Yes", "thinkorswim": "Yes", "TrendSpider": "Best", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Auto-detected trendlines",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "Yes", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Pro"}),
            ("Auto Fibonacci levels",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "Yes", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "No"}),
            ("Candlestick pattern recognition",
             {"TradingView": "Community", "thinkorswim": "No", "TrendSpider": "Best", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "Yes", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Planned"}),
            ("Seasonality heatmap",
             {"TradingView": "Community", "thinkorswim": "No", "TrendSpider": "Yes", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Pro"}),
            ("Custom scripting language",
             {"TradingView": "Pine Script", "thinkorswim": "ThinkScript", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "BQL", "AlphaBreak": "No"}),
            ("Technical opinion signal",
             {"TradingView": "Widget", "thinkorswim": "No", "TrendSpider": "Yes", "Barchart": "Yes",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
        ],
    },
    "Fundamental & Data Features": {
        "header_fill": FUNDAMENTAL_HEADER,
        "note": "These features are dominated by Seeking Alpha, Koyfin, and StockAnalysis. Charting platforms largely ignore them.",
        "features": [
            ("Financial statements (10yr+)",
             {"TradingView": "Basic", "thinkorswim": "5yr", "TrendSpider": "No", "Barchart": "Yes",
              "Seeking Alpha": "Yes", "Finviz": "Snapshot", "Koyfin": "10yr+", "StockAnalysis": "10yr+", "Simply Wall St": "Yes",
              "Bloomberg": "10yr+", "AlphaBreak": "Free"}),
            ("Revenue by segment / geography",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "Yes", "StockAnalysis": "Yes", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Elite"}),
            ("Analyst consensus & targets",
             {"TradingView": "Yes", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "Yes",
              "Seeking Alpha": "Yes", "Finviz": "Yes", "Koyfin": "Yes", "StockAnalysis": "Yes", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Free"}),
            ("Earnings estimates & revisions",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "Yes",
              "Seeking Alpha": "Best", "Finviz": "No", "Koyfin": "Yes", "StockAnalysis": "Yes", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Free"}),
            ("Earnings call transcripts",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "Yes", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Planned"}),
            ("Peer comparison table",
             {"TradingView": "No", "thinkorswim": "Basic", "TrendSpider": "No", "Barchart": "Basic",
              "Seeking Alpha": "Yes", "Finviz": "No", "Koyfin": "Yes", "StockAnalysis": "Limited", "Simply Wall St": "Yes",
              "Bloomberg": "Best", "AlphaBreak": "Pro"}),
            ("Dividend analysis (deep)",
             {"TradingView": "Basic", "thinkorswim": "Basic", "TrendSpider": "No", "Barchart": "Basic",
              "Seeking Alpha": "Best", "Finviz": "Basic", "Koyfin": "Yes", "StockAnalysis": "Yes", "Simply Wall St": "Yes",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Insider trading data",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "Yes",
              "Seeking Alpha": "Yes", "Finviz": "Yes", "Koyfin": "Yes", "StockAnalysis": "Yes", "Simply Wall St": "Yes",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Institutional ownership / 13F",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "Yes",
              "Seeking Alpha": "Yes", "Finviz": "Yes", "Koyfin": "Yes", "StockAnalysis": "Yes", "Simply Wall St": "Yes",
              "Bloomberg": "Yes", "AlphaBreak": "Free"}),
            ("Supply chain mapping",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Elite"}),
            ("DuPont decomposition",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Elite"}),
            ("DCF / intrinsic value model",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "Yes",
              "Bloomberg": "Yes", "AlphaBreak": "Planned"}),
            ("Quant letter grades (A-F)",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "Yes", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "Snowflake",
              "Bloomberg": "No", "AlphaBreak": "Pro"}),
        ],
    },
    "Options Analytics (Charting-Side but Sparse)": {
        "header_fill": CHARTING_HEADER,
        "note": "Options tools skew toward the charting/brokerage side, but even there only thinkorswim and Barchart offer depth. Fundamental platforms ignore options entirely.",
        "features": [
            ("Options chain + Greeks",
             {"TradingView": "No", "thinkorswim": "Best", "TrendSpider": "No", "Barchart": "Yes",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "Basic", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Free"}),
            ("Fair value (Black-Scholes)",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("IV history / vol surface",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "Partial", "Barchart": "Partial",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Elite"}),
            ("Unusual options activity",
             {"TradingView": "No", "thinkorswim": "Sizzle", "TrendSpider": "Partial", "Barchart": "Yes",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Probability of profit",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Strategy builder + P&L diagram",
             {"TradingView": "No", "thinkorswim": "Best", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Planned"}),
            ("Market Maker Move",
             {"TradingView": "No", "thinkorswim": "Yes", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("IV crush modeling",
             {"TradingView": "No", "thinkorswim": "Partial", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Elite"}),
            ("Options flow dashboard",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "Partial", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Elite"}),
        ],
    },
    "Unserved / Blue Ocean (Neither Side)": {
        "header_fill": NEITHER_HEADER,
        "note": "Features that NEITHER the charting platforms NOR the fundamental platforms offer. This is where AlphaBreak differentiates.",
        "features": [
            ("AI trade scoring",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
            ("Composite tech + fundamental score",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
            ("Regime-aware analysis",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
            ("Trade journal",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
            ("AI-scored journal entries",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
            ("Pre-trade plans + post-trade reviews",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Pro"}),
            ("News NLP sentiment scoring",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Pro"}),
            ("Earnings surprise prediction (ML)",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Elite"}),
            ("Signal accuracy tracking",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "No", "Barchart": "No",
              "Seeking Alpha": "Partial", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Free"}),
            ("Dark pool data",
             {"TradingView": "No", "thinkorswim": "No", "TrendSpider": "Partial", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "Yes", "AlphaBreak": "Pro"}),
            ("Smart checklists (multi-indicator)",
             {"TradingView": "No", "thinkorswim": "Custom scan", "TrendSpider": "Yes", "Barchart": "No",
              "Seeking Alpha": "No", "Finviz": "No", "Koyfin": "No", "StockAnalysis": "No", "Simply Wall St": "No",
              "Bloomberg": "No", "AlphaBreak": "Pro"}),
        ],
    },
}

# AlphaBreak tier → green shade
AB_TIER = {"Free": GREEN_FREE, "Pro": GREEN_MID, "Elite": GREEN_PRICEY, "Planned": BLUE}


def status_to_fill(status, comp_name):
    """Return (fill, text, font_color_white)."""
    s = status.strip()
    if s == "No":
        return RED, "No", True
    if s == "Planned":
        return BLUE, "Planned", False
    if s in ("Partial", "Basic", "Snapshot", "Limited", "Community", "Sizzle", "Custom scan", "Snowflake", "Widget"):
        return YELLOW, s, False
    # It's some form of "Yes"
    if comp_name == "AlphaBreak":
        for tier, fill in AB_TIER.items():
            if s == tier:
                return fill, s, False
        return GREEN_FREE, s, False
    else:
        green = COMP_GREENS.get(comp_name, GREEN_MID)
        return green, s, False


def style_cell(cell, fill, text, white_font=False):
    cell.value = text
    cell.fill = fill
    cell.font = Font(name="Calibri", size=10, color="FFFFFF" if white_font else "000000")
    cell.alignment = CENTER
    cell.border = THIN


# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: BIFURCATION MATRIX
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Market Bifurcation"

ALL_COMPS = (
    [c[0] for c in CHARTING_COMPS] +
    [c[0] for c in FUNDAMENTAL_COMPS] +
    [c[0] for c in BRIDGE_COMPS]
)
ALL_PRICES = dict(CHARTING_COMPS + FUNDAMENTAL_COMPS + BRIDGE_COMPS)

# Title
ncols = len(ALL_COMPS) + 1
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
ws.cell(row=1, column=1, value="Market Bifurcation: Charting/TA vs Fundamental/Data").font = BOLD_14
ws.row_dimensions[1].height = 28

# Subtitle row — group labels
row = 2
ws.cell(row=row, column=1, value="").border = THIN
charting_start = 2
charting_end = charting_start + len(CHARTING_COMPS) - 1
fund_start = charting_end + 1
fund_end = fund_start + len(FUNDAMENTAL_COMPS) - 1
bridge_start = fund_end + 1
bridge_end = bridge_start + len(BRIDGE_COMPS) - 1

for c in range(charting_start, charting_end + 1):
    cell = ws.cell(row=row, column=c, value="CHARTING / TA" if c == charting_start else "")
    cell.fill = CHARTING_HEADER
    cell.font = BOLD_WHITE
    cell.alignment = CENTER
    cell.border = THIN
if charting_end > charting_start:
    ws.merge_cells(start_row=row, start_column=charting_start, end_row=row, end_column=charting_end)

for c in range(fund_start, fund_end + 1):
    cell = ws.cell(row=row, column=c, value="FUNDAMENTAL / DATA" if c == fund_start else "")
    cell.fill = FUNDAMENTAL_HEADER
    cell.font = BOLD_WHITE
    cell.alignment = CENTER
    cell.border = THIN
if fund_end > fund_start:
    ws.merge_cells(start_row=row, start_column=fund_start, end_row=row, end_column=fund_end)

for c in range(bridge_start, bridge_end + 1):
    cell = ws.cell(row=row, column=c, value="BRIDGES BOTH SIDES" if c == bridge_start else "")
    cell.fill = BOTH_HEADER
    cell.font = BOLD_WHITE
    cell.alignment = CENTER
    cell.border = THIN
if bridge_end > bridge_start:
    ws.merge_cells(start_row=row, start_column=bridge_start, end_row=row, end_column=bridge_end)

ws.row_dimensions[2].height = 22

# Header row — competitor names + prices
row = 3
feat_cell = ws.cell(row=row, column=1, value="Feature")
feat_cell.fill = HEADER_FILL
feat_cell.font = HEADER_FONT
feat_cell.alignment = CENTER
feat_cell.border = THIN

for col_idx, comp in enumerate(ALL_COMPS, 2):
    cell = ws.cell(row=row, column=col_idx, value=f"{comp}\n{ALL_PRICES[comp]}")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN
ws.row_dimensions[3].height = 40

# Data rows
row = 4
for section_name, section_data in BIFURCATION.items():
    # Section header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    sec_cell = ws.cell(row=row, column=1, value=section_name)
    sec_cell.fill = section_data["header_fill"]
    sec_cell.font = BOLD_12
    sec_cell.alignment = LEFT
    sec_cell.border = THIN
    ws.row_dimensions[row].height = 24
    row += 1

    # Note row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    note_cell = ws.cell(row=row, column=1, value=section_data["note"])
    note_cell.font = ITALIC
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.border = THIN
    ws.row_dimensions[row].height = 30
    row += 1

    # Feature rows
    for feature_name, comp_statuses in section_data["features"]:
        fc = ws.cell(row=row, column=1, value=feature_name)
        fc.font = BOLD
        fc.alignment = LEFT
        fc.border = THIN
        fc.fill = SECTION_FILL

        for col_idx, comp in enumerate(ALL_COMPS, 2):
            status = comp_statuses.get(comp, "No")
            fill, text, white = status_to_fill(status, comp)
            style_cell(ws.cell(row=row, column=col_idx), fill, text, white)

        ws.row_dimensions[row].height = 22
        row += 1

    # Blank spacer
    row += 1

# Column widths
ws.column_dimensions["A"].width = 38
for col_idx in range(2, ncols + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 16

ws.freeze_panes = "B4"


# ═══════════════════════════════════════════════════════════════════════════
# TAB 2: BLOOMBERG vs ALPHABREAK
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Bloomberg vs AlphaBreak")

BB_VS_AB = [
    ("Overview & Fundamentals", [
        ("Company overview / key stats",    "Yes", "Free",    "Parity"),
        ("Financial statements (10yr+)",    "10yr+", "Free",  "AlphaBreak needs deeper history"),
        ("Revenue by segment / geography",  "Yes", "Elite",   "AlphaBreak at $299 vs $2K/mo"),
        ("Analyst consensus & targets",     "Yes", "Free",    "Parity"),
        ("Earnings estimates & revisions",  "Yes", "Free",    "Parity"),
        ("Earnings call transcripts",       "Yes", "Planned", "Bloomberg only shares with SA"),
        ("Peer comparison table",           "Best", "Pro",    "Bloomberg COMP is gold standard"),
        ("Dividend analysis",               "Yes", "Pro",     "Parity at Pro tier"),
        ("Insider trading data",            "Yes", "Pro",     "Parity at Pro tier"),
        ("Institutional ownership / 13F",   "Yes", "Free",    "AlphaBreak: 20 funds, 8.4M rows"),
        ("Supply chain mapping",            "SPLC", "Elite",  "Bloomberg-exclusive feature, AB plans it"),
        ("3-statement financials w/ trends","Yes", "Elite",   "AlphaBreak at 1/80th the price"),
        ("DuPont decomposition",            "Yes", "Elite",   "AlphaBreak at 1/80th the price"),
        ("DCF / intrinsic value model",     "Yes", "Planned", "Bloomberg has full modeling suite"),
    ]),
    ("Charting & Technical Analysis", [
        ("Interactive charting",            "Yes", "Free",    "Parity"),
        ("Technical indicators",            "100+", "Free",   "Bloomberg has more; AB covers essentials"),
        ("Drawing tools",                   "Yes", "Pro",     "Parity at Pro tier"),
        ("Multi-timeframe analysis",        "Yes", "Pro",     "Parity at Pro tier"),
        ("Auto-detected trendlines",        "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Auto Fibonacci levels",           "No",  "No",      "Neither offers"),
        ("Seasonality heatmap",             "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Custom scripting (BQL)",          "BQL", "No",      "ADVANTAGE: Bloomberg"),
    ]),
    ("Options Analytics", [
        ("Options chain + Greeks",          "OMON", "Free",   "Bloomberg more depth; AB free"),
        ("Fair value (Black-Scholes)",      "OVME", "Pro",    "Both offer; Bloomberg included"),
        ("IV history / vol surface",        "SKEW", "Elite",  "Bloomberg best-in-class"),
        ("Unusual options activity",        "Yes", "Pro",     "AB at 1/240th the price"),
        ("Probability of profit",           "Yes", "Pro",     "Parity at Pro tier"),
        ("Strategy builder + P&L diagram",  "Yes", "Planned", "Bloomberg has full suite"),
        ("Market Maker Move",               "Yes", "Pro",     "Parity at Pro tier"),
        ("IV crush modeling",               "Yes", "Elite",   "Parity at Elite tier"),
        ("ORATS historical data",           "No",  "Elite",   "ADVANTAGE: AlphaBreak"),
        ("Options flow dashboard",          "Yes", "Elite",   "AB at 1/80th the price"),
    ]),
    ("AI, Scoring & Quantitative", [
        ("AI trade scoring",                "No",  "Free",    "ADVANTAGE: AlphaBreak"),
        ("Composite tech+fundamental score","No",  "Free",    "ADVANTAGE: AlphaBreak"),
        ("Quant letter grades (A-F)",       "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Smart checklists",                "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Regime-aware analysis",           "No",  "Free",    "ADVANTAGE: AlphaBreak"),
        ("News NLP sentiment scoring",      "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Earnings surprise prediction",    "No",  "Elite",   "ADVANTAGE: AlphaBreak"),
        ("Signal accuracy tracking",        "No",  "Free",    "ADVANTAGE: AlphaBreak"),
        ("Short interest / squeeze score",  "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
    ]),
    ("Portfolio, Journal & Workflow", [
        ("Portfolio tracker",               "PORT", "Free",   "Bloomberg has live trading; AB paper"),
        ("Trade journal",                   "No",  "Free",    "ADVANTAGE: AlphaBreak"),
        ("AI-scored journal entries",       "No",  "Free",    "ADVANTAGE: AlphaBreak"),
        ("Trade thesis builder",            "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Pre/post-trade reviews",          "No",  "Pro",     "ADVANTAGE: AlphaBreak"),
        ("Community / shared journal",      "IB Chat", "Free","Different approach — both social"),
        ("Watchlist",                       "Yes", "Free",    "Parity"),
        ("Email / push notifications",      "Yes", "Free",    "Parity"),
        ("Backtesting",                     "Yes", "Elite",   "Bloomberg more sophisticated"),
        ("Brokerage integration",           "Yes", "Planned", "Bloomberg is the brokerage"),
    ]),
    ("Data & Infrastructure", [
        ("Real-time data",                  "Yes", "Pro",     "Bloomberg included; AB at Pro tier"),
        ("API access",                      "BQL", "API tier","Bloomberg BQL is more powerful"),
        ("WebSocket streaming",             "Yes", "Pro",     "Parity at Pro tier"),
        ("Mobile app",                      "Yes", "Planned", "Bloomberg has mature mobile app"),
        ("Dark pool data",                  "Yes", "Pro",     "Both offer; AB at Pro tier"),
        ("Forex correlations",              "Yes", "Free",    "Both offer; AB 21 pairs, 54yr history"),
    ]),
]

# Title
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws2.cell(row=1, column=1, value="Bloomberg Terminal ($24,000/yr) vs AlphaBreak (Free–$299/mo)").font = BOLD_14
ws2.row_dimensions[1].height = 30

# Subtitle
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws2.cell(row=2, column=1,
    value="Bloomberg: ~$2,000/mo  |  AlphaBreak Elite: $299/mo (85% cheaper)  |  AlphaBreak Pro: $99/mo (95% cheaper)  |  AlphaBreak Free: $0"
).font = Font(name="Calibri", size=11, italic=True, color="444444")
ws2.row_dimensions[2].height = 24

# Headers
row = 3
headers = [
    ("Feature", HEADER_FILL, 38),
    ("Bloomberg", BB_BLUE, 16),
    ("AlphaBreak", AB_GREEN, 16),
    ("Verdict", HEADER_FILL, 42),
]
for col_idx, (hdr, fill, _) in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col_idx, value=hdr)
    cell.fill = fill
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN
ws2.row_dimensions[3].height = 28

# Counters for summary
bb_wins = 0
ab_wins = 0
parity_count = 0

# Data
row = 4
for section_name, features in BB_VS_AB:
    # Section header
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    sc = ws2.cell(row=row, column=1, value=section_name)
    sc.fill = SECTION_FILL
    sc.font = BOLD
    sc.alignment = LEFT
    sc.border = THIN
    ws2.row_dimensions[row].height = 24
    row += 1

    for feat, bb_status, ab_status, verdict in features:
        # Feature name
        fc = ws2.cell(row=row, column=1, value=feat)
        fc.font = BOLD
        fc.alignment = LEFT
        fc.border = THIN

        # Bloomberg column
        bb_fill, bb_text, bb_white = status_to_fill(bb_status, "Bloomberg")
        if bb_status == "No":
            bb_fill = RED
            bb_text = "No"
            bb_white = True
        style_cell(ws2.cell(row=row, column=2), bb_fill, bb_text, bb_white)

        # AlphaBreak column
        ab_fill, ab_text, ab_white = status_to_fill(ab_status, "AlphaBreak")
        if ab_status == "No":
            ab_fill = RED
            ab_text = "No"
            ab_white = True
        style_cell(ws2.cell(row=row, column=3), ab_fill, ab_text, ab_white)

        # Verdict column with conditional formatting
        vc = ws2.cell(row=row, column=4, value=verdict)
        vc.alignment = LEFT
        vc.border = THIN
        if "ADVANTAGE: AlphaBreak" in verdict:
            vc.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            vc.font = Font(name="Calibri", size=10, bold=True, color="375623")
            ab_wins += 1
        elif "ADVANTAGE: Bloomberg" in verdict:
            vc.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            vc.font = Font(name="Calibri", size=10, bold=True, color="833C0C")
            bb_wins += 1
        elif "Parity" in verdict or "Both" in verdict:
            vc.font = NORMAL
            parity_count += 1
        else:
            vc.font = NORMAL

        ws2.row_dimensions[row].height = 22
        row += 1

    row += 1  # spacer

# Summary row
row += 1
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws2.cell(row=row, column=1, value="SCORECARD").font = Font(name="Calibri", bold=True, size=14, color="2F5496")
ws2.row_dimensions[row].height = 28
row += 1

summary_items = [
    ("AlphaBreak Advantages", str(ab_wins), PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")),
    ("Bloomberg Advantages", str(bb_wins), PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")),
    ("Parity / Comparable", str(parity_count), PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")),
    ("Bloomberg Annual Cost", "$24,000", PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")),
    ("AlphaBreak Elite Annual Cost", "$3,588", PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")),
    ("AlphaBreak Pro Annual Cost", "$1,188", PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")),
    ("Savings (Elite vs Bloomberg)", "85%", PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")),
    ("Savings (Pro vs Bloomberg)", "95%", PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")),
]
for label, value, fill in summary_items:
    ws2.cell(row=row, column=1, value=label).font = BOLD
    ws2.cell(row=row, column=1).border = THIN
    ws2.cell(row=row, column=1).fill = fill
    vc = ws2.cell(row=row, column=2, value=value)
    vc.font = Font(name="Calibri", bold=True, size=12)
    vc.alignment = CENTER
    vc.border = THIN
    vc.fill = fill
    ws2.row_dimensions[row].height = 24
    row += 1

# Column widths
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 44

ws2.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════
# TAB 3: LEGEND
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Legend")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 55

legend = [
    ("Color Code", "Meaning", None),
    ("Dark Green", "Free / included with free account", GREEN_FREE),
    ("Medium Green", "Available for <$50/mo", GREEN_CHEAP),
    ("Light Green", "Available for $50-150/mo", GREEN_MID),
    ("Pale Green", "Available for $150+/mo or institutional pricing", GREEN_PRICEY),
    ("Red", "Feature NOT offered", RED),
    ("Yellow", "Partial / basic implementation", YELLOW),
    ("Blue", "Planned for AlphaBreak (not yet built)", BLUE),
    ("", "", None),
    ("Bifurcation Tab Colors", "", None),
    ("Purple header", "Charting / TA-dominant features", CHARTING_HEADER),
    ("Blue header", "Fundamental / data-dominant features", FUNDAMENTAL_HEADER),
    ("Gold header", "Bridges both sides (Bloomberg, AlphaBreak)", BOTH_HEADER),
    ("Gray header", "Neither side serves well (blue ocean for AlphaBreak)", NEITHER_HEADER),
]
for row_idx, (label, desc, fill) in enumerate(legend, 1):
    a = ws3.cell(row=row_idx, column=1, value=label)
    b = ws3.cell(row=row_idx, column=2, value=desc)
    if row_idx == 1 or (label and "Tab" in label):
        a.font = BOLD
        b.font = BOLD
    elif fill:
        a.fill = fill
        needs_white = fill in (RED, CHARTING_HEADER, FUNDAMENTAL_HEADER, BOTH_HEADER, NEITHER_HEADER)
        a.font = Font(name="Calibri", size=11, color="FFFFFF" if needs_white else "000000")


output = r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-SophistryDude\Securities_prediction_model\docs\csv\competitive_bifurcation_analysis.xlsx"
wb.save(output)
print(f"Saved to {output}")
print(f"\nBloomberg vs AlphaBreak Scorecard:")
print(f"  AlphaBreak wins: {ab_wins}")
print(f"  Bloomberg wins:  {bb_wins}")
print(f"  Parity:          {parity_count}")
