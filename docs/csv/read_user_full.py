"""Read ALL columns from user's updated file and compare to original."""
import openpyxl

user_wb = openpyxl.load_workbook(
    r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-SophistryDude\Securities_prediction_model\docs\csv\competitive_feature_matrix_user_input.xlsx",
    data_only=True
)
orig_wb = openpyxl.load_workbook(
    r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-SophistryDude\Securities_prediction_model\docs\csv\competitive_feature_matrix.xlsx",
    data_only=True
)

def get_color(cell):
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb and str(fill.fgColor.rgb) not in ("00000000", "0"):
        return str(fill.fgColor.rgb)
    if fill and fill.start_color and fill.start_color.rgb and str(fill.start_color.rgb) not in ("00000000", "0"):
        return str(fill.start_color.rgb)
    return "none"

for sheet_name in user_wb.sheetnames:
    if sheet_name == "Legend" or sheet_name not in orig_wb.sheetnames:
        continue

    u_ws = user_wb[sheet_name]
    o_ws = orig_wb[sheet_name]

    diffs = []
    for row in range(1, u_ws.max_row + 1):
        for col in range(1, u_ws.max_column + 1):
            u_cell = u_ws.cell(row=row, column=col)
            o_cell = o_ws.cell(row=row, column=col)

            u_val = str(u_cell.value).strip() if u_cell.value is not None else ""
            o_val = str(o_cell.value).strip() if o_cell.value is not None else ""
            u_color = get_color(u_cell)
            o_color = get_color(o_cell)

            if u_val != o_val or u_color != o_color:
                feature = str(u_ws.cell(row=row, column=1).value or "").strip()
                header = str(u_ws.cell(row=2, column=col).value or "").strip()[:20]
                diffs.append({
                    "row": row, "col": col,
                    "feature": feature, "header": header,
                    "orig_val": o_val, "user_val": u_val,
                    "orig_color": o_color, "user_color": u_color,
                })

    if diffs:
        print(f"\n=== {sheet_name} — {len(diffs)} differences ===")
        for d in diffs:
            changes = []
            if d["orig_val"] != d["user_val"]:
                changes.append(f"Value: '{d['orig_val']}' -> '{d['user_val']}'")
            if d["orig_color"] != d["user_color"]:
                changes.append(f"Color: {d['orig_color']} -> {d['user_color']}")
            print(f"  Row {d['row']:3d} Col {d['col']:2d} [{d['header']:20s}] {d['feature'][:40]:40s} | {' | '.join(changes)}")
    else:
        print(f"\n=== {sheet_name} — NO CHANGES ===")
