"""Build competitive analysis Excel file with color-coded feature matrices."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors — different greens for price tiers
GREEN_FREE = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")      # Dark green = free
GREEN_CHEAP = PatternFill(start_color="60C060", end_color="60C060", fill_type="solid")     # Medium green = <$50/mo
GREEN_MID = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")       # Light green = $50-150/mo
GREEN_PRICEY = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Pale green = $150+/mo
RED = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
YELLOW = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Partial
BLUE = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")    # Planned

# Map each competitor to a green shade based on price to access the feature
# Bloomberg=$2000/mo, TradingView=free-$60, thinkorswim=free, SeekingAlpha=free-$42,
# Finviz=free-$40, Koyfin=free-$45, TrendSpider=$32-87, StockAnalysis=free-$10,
# SimplyWallSt=free-$40, Barchart=free-$100, AlphaBreak=free-$299
COMPETITOR_GREEN = {
    0: GREEN_PRICEY,    # Bloomberg - $24K/yr
    1: GREEN_CHEAP,     # TradingView - free-$60/mo
    2: GREEN_FREE,      # thinkorswim - free
    3: GREEN_CHEAP,     # Seeking Alpha - free-$42/mo
    4: GREEN_CHEAP,     # Finviz - free-$40/mo
    5: GREEN_CHEAP,     # Koyfin - free-$45/mo
    6: GREEN_MID,       # TrendSpider - $32-87/mo (no free)
    7: GREEN_FREE,      # StockAnalysis - free-$10/mo
    8: GREEN_CHEAP,     # Simply Wall St - free-$40/mo
    9: GREEN_MID,       # Barchart - free-$100/mo
    10: None,           # AlphaBreak - determined per feature
}

# AlphaBreak tier → green shade
AB_TIER_GREEN = {
    "Free": GREEN_FREE,
    "Pro": GREEN_MID,
    "Elite": GREEN_PRICEY,
    "API tier": GREEN_PRICEY,
}
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=10)
SMALL = Font(name="Calibri", size=9, color="444444")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Competitors with pricing
competitors = [
    ("Bloomberg", "$24,000/yr"),
    ("TradingView", "Free–$60/mo"),
    ("thinkorswim", "Free (w/ acct)"),
    ("Seeking Alpha", "Free–$42/mo"),
    ("Finviz", "Free–$40/mo"),
    ("Koyfin", "Free–$45/mo"),
    ("TrendSpider", "$32–87/mo"),
    ("StockAnalysis", "Free–$10/mo"),
    ("Simply Wall St", "Free–$40/mo"),
    ("Barchart", "Free–$100/mo"),
    ("AlphaBreak", "Free / $99 / $299"),
]

# Feature data: (feature_name, [status per competitor])
# Status: "Y" = yes/green, "N" = no/red, "P" = partial/yellow, "F" = planned/blue, or text note
SECTIONS = {
    "Overview & Fundamentals": [
        ("Company overview / key stats",       ["Y","Y","Y","Y","Y","Y","N","Y","Y","Y","Y Free"]),
        ("Income statement",                    ["Y 10yr+","Y Basic","Y 5yr","Y","Snapshot","Y 10yr+","N","Y 10yr+","Y","Y","Y Free"]),
        ("Balance sheet",                       ["Y 10yr+","Y Basic","Y 5yr","Y","Snapshot","Y 10yr+","N","Y 10yr+","Y","Y","Y Free"]),
        ("Cash flow statement",                 ["Y 10yr+","Y Basic","Y 5yr","Y","Snapshot","Y 10yr+","N","Y 10yr+","Y","Y","Y Free"]),
        ("Revenue by segment",                  ["Y","N","N","N","N","Y","N","Y","N","N","Y Elite"]),
        ("Revenue by geography",                ["Y","N","N","N","N","Y","N","Y","N","N","Y Elite"]),
        ("Analyst consensus & targets",         ["Y","Y","Y","Y","Y","Y","N","Y","N","Y","Y Free"]),
        ("Earnings estimates & revisions",      ["Y","N","Y","Y Best","N","Y","N","Y","N","Y","Y Free"]),
        ("Earnings call transcripts",           ["Y","N","N","Y","N","N","N","N","N","N","F Elite"]),
        ("Peer comparison table",               ["Y Best","N","P Basic","Y","N","Y","N","P Limited","Y","P Basic","Y Pro"]),
        ("Dividend analysis",                   ["Y","P Basic","P Basic","Y Best","P Basic","Y","N","Y","Y","P Basic","Y Pro"]),
        ("Insider trading data",                ["Y","N","Y","Y","Y","Y","N","Y","Y","Y","Y Pro"]),
        ("Institutional ownership / 13F",       ["Y","N","Y","Y","Y","Y","N","Y","Y","Y","Y Free"]),
        ("Supply chain mapping",                ["Y","N","N","N","N","N","N","N","N","N","Y Elite"]),
        ("3-statement financials w/ trends",    ["Y","N","N","N","N","Y","N","Y","N","N","Y Elite"]),
        ("DuPont decomposition",                ["Y","N","N","N","N","N","N","N","N","N","Y Elite"]),
        ("DCF / intrinsic value model",         ["Y","N","N","N","N","N","N","N","Y","N","F Future"]),
    ],
    "Charting & Technical Analysis": [
        ("Interactive charting",                ["Y","Y Best","Y","N","P Basic","Y","Y","N","N","Y","Y Free"]),
        ("Technical indicators (RSI, MACD, etc.)", ["Y 100+","Y 100+","Y 400+","N","P Basic","Y","Y","N","N","Y","Y Free"]),
        ("Drawing tools (Fib, trendlines)",     ["Y","Y Best","Y","N","N","P Basic","Y","N","N","Y","Y Pro"]),
        ("Multi-timeframe analysis",            ["Y","Y","Y","N","N","N","Y Best","N","N","N","Y Pro"]),
        ("Auto-detected trendlines",            ["N","N","N","N","N","N","Y","N","N","N","Y Pro"]),
        ("Auto Fibonacci levels",               ["N","N","N","N","N","N","Y","N","N","N","N"]),
        ("Candlestick pattern recognition",     ["N","P Community","N","N","Y","N","Y Best","N","N","N","F Future"]),
        ("Seasonality patterns (return heatmap)", ["N","P Community","N","N","N","N","Y","N","N","N","Y Pro"]),
        ("Custom scripting language",           ["Y BQL","Y Pine","Y ThinkScript","N","N","N","N","N","N","N","N"]),
        ("Raindrop / volume profile charts",    ["N","P Community","N","N","N","N","Y","N","N","N","F Future"]),
        ("Technical opinion signal",            ["N","Y Widget","N","N","N","N","Y","N","N","Y","Y Free"]),
    ],
    "Options Analytics": [
        ("Options chain",                       ["Y","N","Y","N","N","N","N","P Basic","N","Y","Y Free"]),
        ("Greeks (Delta, Gamma, Theta, Vega)",  ["Y","N","Y Best","N","N","N","N","N","N","Y","Y Elite"]),
        ("Fair value (Black-Scholes, Binomial)",["Y","N","Y","N","N","N","N","N","N","N","Y Pro"]),
        ("IV history / vol surface",            ["Y","N","Y","N","N","N","P","N","N","P","Y Elite"]),
        ("Unusual options activity",            ["Y","N","Y Sizzle","N","N","N","P","N","N","Y","Y Pro"]),
        ("Probability of profit",               ["Y","N","Y","N","N","N","N","N","N","N","Y Pro"]),
        ("Risk profile / P&L diagram",          ["Y","N","Y","N","N","N","N","N","N","N","F Future"]),
        ("Strategy builder (spreads, condors)", ["Y","N","Y Best","N","N","N","N","N","N","N","F Future"]),
        ("Market Maker Move",                   ["Y","N","Y","N","N","N","N","N","N","N","Y Pro"]),
        ("IV crush modeling",                   ["Y","N","P","N","N","N","N","N","N","N","Y Elite"]),
        ("ORATS historical data",               ["N","N","N","N","N","N","N","N","N","N","Y Elite"]),
        ("Options flow dashboard",              ["Y","N","N","N","N","N","P","N","N","N","Y Elite"]),
    ],
    "AI, Scoring & Quantitative": [
        ("Quant letter grades (A-F)",           ["N","N","N","Y","N","N","N","N","P Snowflake","N","Y Pro"]),
        ("AI trade scoring",                    ["N","N","N","N","N","N","N","N","N","N","Y Free"]),
        ("Composite tech + fundamental score",  ["N","N","N","N","N","N","N","N","N","N","Y Free"]),
        ("Smart checklists (multi-indicator)",   ["N","N","P Custom scan","N","N","N","Y","N","N","N","Y Pro"]),
        ("Regime-aware analysis",               ["N","N","N","N","N","N","N","N","N","N","Y Free"]),
        ("News NLP sentiment scoring",          ["N","N","N","N","N","N","N","N","N","N","Y Pro"]),
        ("Earnings surprise prediction (ML)",   ["N","N","N","N","N","N","N","N","N","N","Y Elite"]),
        ("Signal accuracy tracking",            ["N","N","N","P","N","N","N","N","N","N","Y Free"]),
        ("Short interest / squeeze score",      ["N","N","N","N","Y","N","N","N","N","N","Y Pro"]),
    ],
    "Portfolio, Journal & Workflow": [
        ("Portfolio tracker",                   ["Y","N","Y Live","Y Basic","N","Y","N","N","Y","N","Y Free"]),
        ("Trade journal",                       ["N","N","N","N","N","N","N","N","N","N","Y Free"]),
        ("AI-scored journal entries",           ["N","N","N","N","N","N","N","N","N","N","Y Free"]),
        ("Trade thesis builder",                ["N","N","N","N","N","N","N","N","N","N","Y Pro"]),
        ("Pre-trade plans + post-trade reviews",["N","N","N","N","N","N","N","N","N","N","Y Pro"]),
        ("Community / shared journal",          ["N","Y Ideas","N","Y Articles","N","N","N","N","N","N","Y Free"]),
        ("Watchlist",                           ["Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y Free"]),
        ("Email / push notifications",          ["Y","Y","Y","Y","N","N","Y","N","N","Y","Y Free"]),
        ("Backtesting",                         ["Y","Y Pine","Y","N","N","N","Y","N","N","N","Y Elite"]),
        ("Brokerage integration",               ["Y","Y Multi","Y Schwab","N","N","N","N","N","N","N","F Q4"]),
    ],
    "Data & Infrastructure": [
        ("Real-time data",                      ["Y","Y Paid","Y","N","Y Paid","N","Y Paid","N","N","Y Paid","Y Pro"]),
        ("Delayed data (free)",                 ["N","Y","Y","Y","Y","Y","N","Y","Y","Y","Y Free"]),
        ("API access",                          ["Y BQL","N","N","N","N","N","N","N","N","N","Y API tier"]),
        ("WebSocket streaming",                 ["Y","Y","Y","N","N","N","N","N","N","N","Y Pro"]),
        ("Mobile app",                          ["Y","Y","Y","Y","N","Y","Y","N","Y","Y","F Q3"]),
        ("Dark pool data",                      ["Y","N","N","N","N","N","P","N","N","N","Y Pro"]),
        ("Forex correlations",                  ["Y","Y","Y","N","N","N","N","N","N","N","Y Free"]),
    ],
}


def get_fill_and_text(status, comp_idx):
    """Return (fill, display_text) based on status string and competitor index."""
    s = status.strip()
    if s.startswith("F"):
        # Planned/future (AlphaBreak only)
        label = s[1:].strip() if len(s) > 1 else "Planned"
        return BLUE, f"Planned ({label})" if label else "Planned"
    elif s.startswith("Y"):
        label = s[1:].strip()
        # Determine green shade
        if comp_idx == 10:  # AlphaBreak
            # Parse tier from label
            for tier, fill in AB_TIER_GREEN.items():
                if tier in label:
                    display = label
                    return fill, display
            return GREEN_FREE, label if label else "Yes"
        else:
            green = COMPETITOR_GREEN.get(comp_idx, GREEN_MID)
            return green, label if label else "Yes"
    elif s.startswith("P"):
        label = s[1:].strip()
        return YELLOW, label if label else "Partial"
    elif s == "N":
        return RED, "No"
    elif s.startswith("Snapshot"):
        return YELLOW, "Snapshot"
    else:
        return YELLOW, s


def write_sheet(ws, section_name, features):
    """Write one section as a sheet."""
    ws.title = section_name[:31]  # Excel tab name limit

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(competitors) + 1)
    title_cell = ws.cell(row=1, column=1, value=section_name)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2: Headers
    headers = ["Feature"] + [f"{name}\n{price}" for name, price in competitors]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 40

    # Data rows
    for row_idx, (feature, statuses) in enumerate(features, 3):
        # Feature name
        feat_cell = ws.cell(row=row_idx, column=1, value=feature)
        feat_cell.font = BOLD
        feat_cell.alignment = LEFT
        feat_cell.border = THIN_BORDER
        feat_cell.fill = SUBHEADER_FILL

        # Competitor columns
        for col_idx, status in enumerate(statuses, 2):
            comp_idx = col_idx - 2  # 0=Bloomberg, 1=TradingView, ..., 10=AlphaBreak
            fill, text = get_fill_and_text(status, comp_idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.fill = fill
            cell.font = NORMAL if fill != RED else Font(name="Calibri", size=10, color="FFFFFF")
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_idx].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 38
    for col_idx in range(2, len(competitors) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # Freeze header row
    ws.freeze_panes = "B3"


# Create sheets
first = True
for section_name, features in SECTIONS.items():
    if first:
        ws = wb.active
        first = False
    else:
        ws = wb.create_sheet()
    write_sheet(ws, section_name, features)

# Add Legend sheet
ws_legend = wb.create_sheet("Legend")
ws_legend.column_dimensions["A"].width = 20
ws_legend.column_dimensions["B"].width = 50

legend_items = [
    ("Color Code", "Meaning", None),
    ("Dark Green", "Free / included with free account", GREEN_FREE),
    ("Medium Green", "Available for <$50/mo", GREEN_CHEAP),
    ("Light Green", "Available for $50–150/mo", GREEN_MID),
    ("Pale Green", "Available for $150+/mo or institutional pricing", GREEN_PRICEY),
    ("Red", "Feature NOT offered", RED),
    ("Yellow", "Partial / basic implementation", YELLOW),
    ("Blue", "Planned for AlphaBreak (not yet built)", BLUE),
]
for row_idx, (label, desc, fill) in enumerate(legend_items, 1):
    a = ws_legend.cell(row=row_idx, column=1, value=label)
    b = ws_legend.cell(row=row_idx, column=2, value=desc)
    if row_idx == 1:
        a.font = BOLD
        b.font = BOLD
    elif fill:
        a.fill = fill
        a.font = Font(name="Calibri", size=11, color="FFFFFF" if label == "Red" else "000000")

output = r"c:\Users\nicho\OneDrive\Documents\GitHub\data-acq-functional-SophistryDude\AlphaBreak\docs\csv\competitive_feature_matrix.xlsx"
wb.save(output)
print(f"Saved to {output}")
